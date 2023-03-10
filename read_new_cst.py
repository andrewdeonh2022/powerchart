import openpyxl
from xml.dom import minidom
from collections import defaultdict
import glob
import json

# check for map entry errors
errorlist = ["calcium ionized whole blood cwh", "hiv 1/2 antibody and p24 antigen bccdc", "hiv 1/2 antibody and p24 antigen phc", "holo-transcobalamin ii active b12", "influenza a, b and rsv nat bccdc", "influenza a/b and rsv nat phc", "influenza a/b nat cwh", "mri cardiac w/ contrast", "respiratory syncytial virus (rsv) nat cwh", "serum creatinine, alt  (if no recent results in past 3 months)", "dextrose 5% (d5w) intermittent flush", "eplerenone", "ferritin", "ranitidine"]

def GetCurrentVersion():

	tversion = 0
	with open("version_code.tsv", "r") as infile:
		for line in infile:
			line = line.strip()
			ldata = line.split("\t")
			if ldata[0] == "version":
				tversion = round(float(ldata[1]), 2)
	infile.close()

	nversion = tversion + 0.01
	nversion = round(nversion, 2)
	f = open("version_code.tsv", "w")
	f.write("version\t" + str(nversion))
	f.close()
	
	return tversion

def LoadOrderSetIndex():

	index_map = defaultdict(int)
	with open("order_set_index.tsv", "r") as infile:
		for line in infile:
			line = line.strip()
			ldata = line.split("\t")
			tfile = ldata[1]
			tindex = ldata[0]
			
	infile.close()
	
	return index_map

def LoadCurrentMapping():

	current_map = defaultdict(str)
	#with open("mapping_mar3_2023.tsv", "r", encoding="utf-8") as infile:
	#with open("mapping_mar8_2023.tsv", "r", encoding="utf-8") as infile:
	with open("mapping_mar9_2023.tsv", "r", encoding="utf-8") as infile:
		theader = infile.readline()
		for line in infile:
		
			line = line.strip()
			ldata = line.split("\t")

			if len(ldata) > 2:
				tbase = ldata[0].lower().strip()
				tbase = tbase.replace("stat", "")
				tbase = ' '.join(tbase.split())
				tval = ldata[1]
				if tval.lower().strip() == "insert note":
					tval = "insert note::" + ldata[3]
				current_map[tbase] = tval

				# remove stat
				tbase = tbase.replace("stat", "")
				tbase = ' '.join(tbase.split())
				current_map[tbase] = tval
				
				# check
				if tbase.find("electrocardiogram") > -1:
					print(tbase + "\t" + tval)
			
	infile.close()

	return current_map

	# Use colors from excel file
	# https://stackoverflow.com/questions/32736419/get-cell-color-from-xlsx
	#book = openpyxl.load_workbook(tfile)
	#sheet = book.active 
	#row_count = sheet.max_row
	#column_count = sheet.max_column
	#matchdata = defaultdict(str)
	#for x in range(1, row_count):
	#	# match on e, f
	#	val1 = str(sheet.cell(row=x, column=5).value).strip().lower()
	#	val2 = str(sheet.cell(row=x, column=6).value).strip().lower()
	#	color1 = sheet.cell(row=x, column=5).fill.start_color.index
	#	color2 = sheet.cell(row=x, column=6).fill.start_color.index
	#	
	#	#print("NH data: " + "\t" + str(val1) + "\t" + str(val2))
	#	matchdata[val1] = ""
	#	matchdata[val2] = ""

	#return matchdata

def GetCSTList(cst_path):

	#cst_path = "c:\\Apache24\\htdocs\\orders\\data\\high_priority\\DCWs"

	#subdirs = []
	#for tfile in glob.glob(cst_path + "\\*"):
	#	if ((tfile.find('.zip') == -1) and (tfile.find('.xlsm') == -1)):
	#		subdirs.append(tfile)

	filelist = []
	for tfile in glob.glob(cst_path + "\\*.xlsx"):
		filelist.append(tfile)

	tfile = "high_priority.tsv"
	f = open(tfile, "w")
	for x in range(0, len(filelist)):
		f.write(str(x) + "\t" + filelist[x] + "\n")
	f.close()
	
	return filelist

def LoadNHOrderData():

	tfile = "Copy of NIRH_CD Order Catalog - Complete (30 NOV 2022) - Unfiltered.xlsx"
	
	#tfile = "DCW Library - All PowerPlans in P0783 (DCW Format).xlsx"
	book = openpyxl.load_workbook(tfile)
	sheet = book.active 
	row_count = sheet.max_row
	column_count = sheet.max_column

	matchdata = defaultdict(str)
	for x in range(1, row_count):
		# match on e, f
		val1 = str(sheet.cell(row=x, column=5).value)
		val2 = str(sheet.cell(row=x, column=6).value)
		oval1 = val1.lower().strip()
		oval2 = val2.lower().strip()
		
		#print("NH data: " + "\t" + str(val1) + "\t" + str(val2))
		matchdata[oval1] = val1
		matchdata[oval2] = val2
		matchdata[val1] = val1
		matchdata[val2] = val2

	return matchdata

def LoadNHToCSTMap():

	# updated from OrderSetMapping.xlsx
	tablemap = defaultdict(str)
	with open("nh_to_cst_revised_map.tsv", "r") as infile:
		theader = infile.readline()
		for line in infile:
			line = line.strip()
			ldata = line.split("\t")
			tcst = ldata[0].lower().strip()
			tnh = ldata[1].strip()
			tablemap[tcst] = tnh
	infile.close()

	return tablemap

#def MatchToNHOrders(tval, matchdata, match_results):
def MatchToNHOrders(tval, matchdata, match_results):

	nval = tval.strip().lower()
	if nval in matchdata:
		match_results[nval] = nval
		#print("Found match in NH data: " + nval)
	if nval not in matchdata:
		match_results[nval] = "No NH match"
		#print("MISSING Found in CST Only: " + nval)
	
	return match_results
	
def LoadPriorityCSTList(tfile):

	filelist = []
	#tfile = "high_priority.tsv"
	with open(tfile, "r") as infile:
		for line in infile:
			line = line.strip()
			ldata = line.split("\t")
			filelist.append(ldata[1])
	infile.close()
	
	return filelist
	
def PrintLine(trow):

	tline = ""
	for x in range(0, len(trow)):
		tline += trow[x] + "\t"
	tline = tline.strip() + "\n"
	
	return tline
	
def GetIndent(tlevel):

	tline = ""
	for x in range(0, tlevel):
		tline += "\t"
	
	return tline

def PrintHTMLLine(trow):

	tline = "<tr>"
	for x in range(0, len(trow)):
		tline += "<td>" + trow[x] + "</td>"
	tline = tline.strip() + "</tr>\n"
	
	return tline


def MakeHTMLFile(tid, ttitle, collist, ttable):

	tfile = str(tid) + ".html"
	f = open(tfile, "w")
	f.write("<html><head><title>" + ttitle + " (Version " + str() + ")</title>")
	f.write('<link rel="stylesheet" href="cst_table.css">' + "\n")
	f.write("</head><body style=\"font-family: arial;\">")
	f.write("<h2>" + ttitle + "</h2>\n")
	f.write("<table id=\"customers\">\n")
	theader = PrintHTMLLine(collist)
	f.write(theader)
	for trow in ttable:
		tline = PrintHTMLLine(trow)
		f.write(tline)
	f.write("</table></body></html>\n")
	f.close()

def MakeJSONIndex(datatable):

	tjson = '[\n'
	for tid in datatable:
		tjson += "\t" + '{"dcw_id": "' + tid + '", "title": "' + datatable[tid] + '"},' + "\n"
	tjson = tjson[:-2] + "\n"
	tjson += ']\n'

	f = open("cst_dcws.json", "w")
	f.write(tjson)
	f.close()

def GetOrderSetProp():

	txml = ""
	with open("head_xml_data.dat", "r") as infile:
		for line in infile:
			line = line.strip()
			txml = txml + "\t\t" + line + "\n"
	infile.close()
	
	return txml

def EncodeURL(tval):

	tval = tval.replace('&', '&amp;')
	
	return tval

def FormatStr(tfield):

	#if (tfield.find("http:") > -1) or (tfield.find("https:") > -1) or (tfield.find("www.") > -1):
	nfield = tfield.replace('&', '&amp;')
	nfield = nfield.replace("\n", "&#10;&#13;")
	nfield = nfield.replace("<", "&lt;")
	nfield = nfield.replace(">", "&gt;")
	nfield = nfield.replace('???', '"')
	nfield = nfield.replace('???', '"')

	return nfield

def FormatTitle(tfield):

	nfield = tfield.replace('&', 'and')
	nfield = nfield.replace('???', '"')
	nfield = nfield.replace('???', '"')

	return nfield

def CreateHTMLPages():

	tfile = "DCW Library - All PowerPlans in P0783 (DCW Format).xlsx"
	book = openpyxl.load_workbook(tfile)
	sheet = book.active 
	row_count = sheet.max_row
	column_count = sheet.max_column

	orderdata = defaultdict(lambda: defaultdict(str))
	tplan_name = ""
	tphase = ""
	tcat = ""
	tsubcat = ""
	ttype = ""
	tcnt = 1
	orderdata = {}	
	for x in range(1, row_count):
		tval = sheet.cell(row=x, column=1).value
		if tval != "":
			orderdata[tplan_name]["name"] = tval	

def GetNHCategory(tval):

	# NH Categories on B1978 from Tyler Scott (Dec 20, 2022)
	# NH Category                   CST Category
	# -----------                   ------------
	# Admit/Transfer/Discharge      Admit/Transfer/Discharge
	# Patient Care                  Patient Care
	# Activity                      Activity
	# Diet/Nutrition                Diet/Nutrition
	# Continuous Infusions          Continuous Infusions
	# Medications                   Medications
	# Laboratory                    Laboratory
	# Diagnostic Tests              Diagnostic Tests
	# Respiratory                   Respiratory
	# Allied Health                 Allied Health
	# Consults/Referrals            Consults/Referrals
	# Communication Orders          Communication Orders
	# Procedures                    Procedures
	# Non Categorized               -----
	# Allergies                     -----
	# Diagnoses                     -----
	# Medical Supplies              Supplies
	# Special Procedures            -----
	# Other Test                    -----
	# Supplies                      Supplies
	
	nval = tval
	if tval == "Allergies":
		nval = "Non Categorized"
	if tval == "Diagnoses":
		nval = "Non Categorized"
	if tval == "Medical Supplies":
		nval = "Non Categorized"
	if tval == "Special Procedures":
		nval = "Non Categorized"
	if tval == "Other Test":
		nval = "Non Categorized"
	
	
	retval = nval + "|Clinical Categories"

	return retval

def XMLBuildDetailList(dcwid, catid, compid, sentid, detail_list):

	# count up details
	tXML = ""
	tcnt = 0
	for tid in detail_list[dcwid][catid][compid][sentid]:
		tcnt += 1
	
	if tcnt > 0:
		tXML = "\n" + GetIndent(9) + "<DETAILLIST>\n"
		for tid in detail_list[dcwid][catid][compid][sentid]:
			tfield = str(detail_list[dcwid][catid][compid][sentid][tid])
			# format field carriage returns
			tfield = FormatStr(tfield)
			tXML += GetIndent(10) + "<DETAILS>\n"
			tXML += GetIndent(11) + "<FIELDMEAN>" + tid + "</FIELDMEAN>\n"
			tXML += GetIndent(11) + "<FIELDDESC>" + tid + "</FIELDDESC>\n"
			tXML += GetIndent(11) + "<FIELDDISPVALUE>" + tfield + "</FIELDDISPVALUE>\n"
			tXML += GetIndent(10) + "</DETAILS>\n"
		tXML += GetIndent(9) + "</DETAILLIST>\n"

	return tXML

def XMLBuildSentenceList(dcwid, catid, compid, sentence_list, detail_list):

	# count up sentences
	tXML = ""
	tcnt = 0
	for tid in sentence_list[dcwid][catid][compid]:
		tcnt += 1
	
	if tcnt > 0:
		tXML = "\n" + GetIndent(7) + "<SENTENCELIST>\n"
		for tid in sentence_list[dcwid][catid][compid]:
			tsentence = sentence_list[dcwid][catid][compid][tid]
			tsentence = FormatStr(tsentence)
			if (tsentence != "None") and (tsentence != ""):
				tXML += GetIndent(8) + "<SENTENCE>\n"
				tXML += GetIndent(9) + "<DISPLAYLINE>" + tsentence + "</DISPLAYLINE>\n"
				#tXML += GetIndent(9) + tsentence
				tXML += "\n" + GetIndent(8) + "</SENTENCE>\n"
		tXML += GetIndent(7) + "</SENTENCELIST>\n"

	return tXML
	
def XMLBuildComponentList(dcwid, catid, component_list, sentence_list, detail_list, matchdata, match_results, ordermap, component_type):

	showdebug = 0
	tXML = "\n" + GetIndent(5) + "<COMPONENTLIST>\n"
	numOrders = 0
	numNotes = 0
	for tid in component_list[dcwid][catid]:
	
		# check if note or order - component_type[str(dcwID)][currCategory][tcomponent] = tcompType
		tcomponent_type = component_type[dcwid][catid][tid].lower().strip()
		tcaption = component_list[dcwid][catid][tid]
		tcaption = FormatStr(tcaption)
		showdebug = 0

		#if showdebug == 1:
		#print("tcaption value: " + tcaption)
		#print("type: " + tcomponent_type)
		
		ttype = "O"
		if tcomponent_type == "note":	
			ttype = "L"
			
		nXML = ''
		if (tcaption != "") and (tcaption != "None"):
			
			nXML += GetIndent(6) + "<COMPONENT>\n"
			ncaption = tcaption.lower().strip()
			ncaption = ' '.join(ncaption.split())
			pcaption = FormatStr(tcaption).strip()
			porder = ordermap[ncaption].lower().strip()
			tdup = False
			if tcaption.find('duplicate') > -1:
				tdup = True
				tindex = ncaption.index('duplicate')
				dupsuffix = ncaption[tindex:]
				ncaption = ncaption[0:tindex].lower().strip()
				pcaption = tcaption
			
			# check if order found in order map
			tfound = False
			ncaption = ncaption.replace('stat', "")
			ncaption = ' '.join(ncaption.split())
			if ncaption in ordermap:
				porder = ordermap[ncaption].lower().strip()
				tfound = True
				
			#if tfound == False:
			#	if ncaption in ordermap:
			#		tfound = True
			#		porder = ordermap[ncaption].lower().strip()
			#		print("Order: " + porder + "\t" + nxcaption)

			#print("ncaption: " + ncaption)
								
			if (porder == "") and (tcomponent_type != "note"):
				#if tfound == False:
				match_results[pcaption] = "No NH match"

			nhfound = False			
			# check if the order is found in NH order catalog
			if ncaption in matchdata:
				nhfound = True
				match_results[pcaption] = "Exact Match"
				
			if tfound == True:
			
				# map orders
				if (porder != 'no nh match') and (porder != 'ignore') and (porder.find('insert note::') == -1):
					pcaption = FormatStr(ordermap[ncaption])			
			
				if (porder == 'no nh match'):
					pcaption = "NH Order to be configured: " + pcaption
					ttype = "L"
					
				if (porder == 'ignore'):
					pcaption = "ignore"
					ttype = "X"				
					
				if (porder.find('insert note::') > -1):
					pcaption = ordermap[ncaption].replace('insert note::', '')
					pcaption = FormatStr(pcaption).strip()
					ttype = "L"

			if tdup == True:
				pcaption = "Duplicate order to be configured: " + tcaption
				ttype = "L"
							
			if ncaption == "ignore":
				ttype = "X"
			
			# bug fix - if it shows up as empty...
			if pcaption == "":
				pcaption = tcaption

			#print("ncaption: " + ncaption + "\t" + " pcaption: " + pcaption)
			
			if (tcomponent_type == "note") or (ttype == "L") or (ncaption not in ordermap):
				if ncaption != 'ignore':
					nXML += GetIndent(7) + "<CAPTION>" + pcaption + "</CAPTION>"
					ttype = "L"
			
			if ttype == "O":
				# check to see if caption exists in NH orders:
				#match_results = MatchToNHOrders(tcaption, matchdata, match_results)
				#match_results = MatchToNHOrders(ncaption, matchdata, match_results)

				# check for errors
				#if pcaption in errorlist:
				#	print("Error found: " + pcaption)

				# record missing matches
				if tfound == False:
					match_results[tcaption] = "No NH match"

				nXML += GetIndent(7) + "<CAPTION>" + pcaption + "</CAPTION>\n"			
			
			nXML += "\n" + GetIndent(7) + "<CKI/>\n"
			nXML += GetIndent(7) + "<DEFAULTOSIND>1</DEFAULTOSIND>\n"
			nXML += GetIndent(7) + "<DEFAULTSELECTED>1</DEFAULTSELECTED>\n"
			nXML += GetIndent(7) + "<ORDERABLETYPE>" + ttype + "</ORDERABLETYPE>\n"
			#print(dcwid + " - Sentences: " + str(sentence_list[dcwid][catid][tid]))
			#tXML += GetIndent(7) + XMLBuildSentenceList(dcwid, catid, tid, sentence_list, detail_list)
			nXML += GetIndent(6) + "</COMPONENT>\n"
			
			if (ttype != 'X') and (pcaption != 'ignore') and (pcaption.find('ignore') == -1):
				tXML += nXML
				
				# count up
				#if ttype == 'L':
				#	numNotes += 1
				#if ttype == 'O':
				#	numOrders += 1
				
	if (showdebug == 1):
		print(nXML)
			
	tXML += GetIndent(5) + "</COMPONENTLIST>\n"
	
	#print("Number of notes: " + str(numNotes))
	#print("Number of orders: " + str(numOrders))

	return tXML, match_results

def XMLBuildCategoryList(dcwid, category_list, component_list, sentence_list, detail_list, matchdata, match_results, ordermap, component_type):

	tXML = GetIndent(3) + "<CATEGORYLIST>\n"
	for tid in category_list[dcwid]:
	
		# get NH category
		ncat = GetNHCategory(tid)
		xmlstr, match_results = XMLBuildComponentList(dcwid, tid, component_list, sentence_list, detail_list, matchdata, match_results, ordermap, component_type)
	
		tXML += GetIndent(4) + "<CATEGORY>\n"
		tXML += GetIndent(5) + "<CAPTION>" + tid + "</CAPTION>\n"
		tXML += GetIndent(5) + "<CLINICALCATEGORYMEAN>" + ncat + "</CLINICALCATEGORYMEAN>\n"
		#tXML += GetIndent(3) + "<meaning>" + category_list[tid] + "</meaning>\n"
		tXML += GetIndent(3) + xmlstr
		tXML += GetIndent(4) + "</CATEGORY>\n"
	tXML += GetIndent(3) + "</CATEGORYLIST>\n"

	return tXML, match_results

def XMLMakeRoot(ttitle, tphase):

	ntitle = FormatTitle(ttitle)
	tXML = '<?xml version="1.0" encoding="UTF-8"?>' + "\n"
	tXML += "<KNOWLEDGEPLAN>\n"
	tXML += "\t" + "<SOURCEORDERSETTYPE>CAREPLAN</SOURCEORDERSETTYPE>\n"
	tXML += "\t" + "<CAPTION>" + ntitle + " (Version " + str(mainversion) + ")</CAPTION>\n"
	tXML += "\t" + "<ORDERSETLIST>\n"
	tXML += "\t\t" + "<ORDERSET>\n"
	tXML += "\t\t\t" + "<CAPTION>" + ntitle + " (Version " + str(mainversion) + ")</CAPTION>\n"
	tXML += "\t\t\t" + "<DISPLAY>" + ntitle + " (Version " + str(mainversion) + ")</DISPLAY>\n"
	tXML += "\t\t\t" + "<EVIDENCEURL/>" + "\n"
	tXML += "\t\t\t" + "<EVIDENCEURLTYPEMEAN/>" + "\n"
	tXML += "\t\t\t" + "<CKI/>" + "\n"
	# don't change this version field - it'a a binary flag
	tXML += "\t\t\t" + "<VERSION>1</VERSION>" + "\n"
	tXML += "\t\t\t" + "<DURATION>0</DURATION>" + "\n"
	tXML += "\t\t\t" + "<DURATIONUNITMEAN/>" + "\n"
	tXML += "\t\t\t" + "<SUBPHASEIND>0</SUBPHASEIND>" + "\n"
	tXML += "\t\t\t" + "<STANDARDCYCLENBR>0</STANDARDCYCLENBR>" + "\n"
	tXML += "\t\t\t" + "<CYCLEIND>0</CYCLEIND>" + "\n"
	tXML += "\t\t\t" + "<DIAGNOSISCAPTUREIND>0</DIAGNOSISCAPTUREIND>" + "\n"
	tXML += "\t\t\t" + "<PROVIDERPROMPTIND>0</PROVIDERPROMPTIND>" + "\n"
	tXML += "\t\t\t" + "<HIDEFLEXEDCOMPIND>0</HIDEFLEXEDCOMPIND>" + "\n"
	tXML += "\t\t\t" + "<DEFAULTVIEWMEAN/>" + "\n"
	tXML += "\t\t\t" + "<ALLOWCOPYFORWARDIND>0</ALLOWCOPYFORWARDIND>" + "\n"
	tXML += "\t\t\t" + "<AUTOINITIATEIND>0</AUTOINITIATEIND>" + "\n"
	tXML += "\t\t\t" + "<ALERTSONPLANIND>0</ALERTSONPLANIND>" + "\n"
	tXML += "\t\t\t" + "<ALERTSONPLANUPDIND>0</ALERTSONPLANUPDIND>" + "\n"
	tXML += "\t\t\t" + "<PROMPTONSELECTIONIND>0</PROMPTONSELECTIONIND>" + "\n"
	tXML += "\t\t\t" + "<DEFAULTVISITTYPEFLAG>0</DEFAULTVISITTYPEFLAG>" + "\n"
	tXML += "\t\t\t" + "<CYCLEBEGINNBR>0</CYCLEBEGINNBR>" + "\n"
	tXML += "\t\t\t" + "<CYCLEENDNBR>0</CYCLEENDNBR>" + "\n"
	tXML += "\t\t\t" + "<CYCLELABELDISP/>" + "\n"
	tXML += "\t\t\t" + "<CYCLELOCKENDIND>0</CYCLELOCKENDIND>" + "\n"
	tXML += "\t\t\t" + "<CYCLEDISPLAYENDIND>0</CYCLEDISPLAYENDIND>" + "\n"
	tXML += "\t\t\t" + "<CYCLEINCREMENTNBR>0</CYCLEINCREMENTNBR>" + "\n"
	tXML += "\t\t\t" + "<DEFAULTACTIONINFUTUREMEAN/>" + "\n"
	tXML += "\t\t\t" + "<DEFAULTACTIONINNOWMEAN/>" + "\n"
	tXML += "\t\t\t" + "<DEFAULTACTIONOUTFUTUREMEAN/>" + "\n"
	tXML += "\t\t\t" + "<DEFAULTACTIONOUTNOWMEAN/>" + "\n"
	tXML += "\t\t\t" + "<FUTUREIND>0</FUTUREIND>" + "\n"
	tXML += "\t\t\t" + "<OPTIONALIND>0</OPTIONALIND>" + "\n"
	tXML += "\t\t\t" + "<PATHWAYCLASSMEAN/>" + "\n"
	tXML += "\t\t\t" + "<ROUTEFORREVIEWIND>0</ROUTEFORREVIEWIND>" + "\n"
	tXML += "\t\t\t" + "<DEFAULTSTARTTIMETXT/>" + "\n"
	tXML += "\t\t\t" + "<PRIMARYIND>0</PRIMARYIND>" + "\n"
	tXML += "\t\t\t" + "<PERIODNBR>0</PERIODNBR>" + "\n"
	tXML += "\t\t\t" + "<PERIODCUSTOMLABEL/>" + "\n"
	tXML += "\t\t\t" + "<RESCHEDULEREASONACCEPTFLAG>0</RESCHEDULEREASONACCEPTFLAG>" + "\n"
	tXML += "\t\t\t" + "<OPENBYDEFAULTIND>0</OPENBYDEFAULTIND>" + "\n"
	tXML += "\t\t\t" + "<ALLOWACTIVATEALLIND>1</ALLOWACTIVATEALLIND>" + "\n"
	tXML += "\t\t\t" + "<REVIEWREQUIREDSIGCOUNT>0</REVIEWREQUIREDSIGCOUNT>" + "\n"
	tXML += "\t\t\t" + "<RESTRICTEDACTIONSBITMASK>0</RESTRICTEDACTIONSBITMASK>" + "\n"
	tXML += "\t\t\t" + "<OVERRIDEMRDONPLANIND>0</OVERRIDEMRDONPLANIND>" + "\n"
	tXML += "\t\t\t" + "<LINKEDPHASEIND>0</LINKEDPHASEIND>" + "\n"

	tXML += '{%CategoryList%}'

	tXML += "\n\t\t" + "</ORDERSET>\n"	
	tXML += "\t" + "</ORDERSETLIST>\n"
	tXML += "</KNOWLEDGEPLAN>\n"

	return tXML
	
def XMLMakeFiles(tfile, dcw_index, phaselist, categorylist, component_type, comp_precheck, comp_required, componentlist, sentencelist, detaillist, matchdata, match_results, ordermap, file_index):

	# Create the XML Document Stub and Components
	# Using C:\Apache24\htdocs\orders\data\island_xml\Subcutaneous Insulin Cerner PowePlan V3.xml
	
	opath = "C:\\Apache24\\htdocs\\orders\\data\\high_priority\\XML"
	
	for tid in dcw_index:
	
		#print("Making XML file: " + tid)
		ttitle = dcw_index[tid]
		tphase = phaselist[tid]
		
		# Create XML Root Code
		MainXML = XMLMakeRoot(ttitle, tphase)
		
		# Create Category Data
		CategoryXML, match_results = XMLBuildCategoryList(tid, categorylist, componentlist, sentencelist, detaillist, matchdata, match_results, ordermap, component_type)
		MainXML = MainXML.replace('{%CategoryList%}', CategoryXML)

		# Get Filename for XML
		tdata = tfile.split('\\')
		ofile = tdata[len(tdata) - 1]
		
		tnfile = opath + "\\" + str(file_index) + " - " + ofile + ".xml"
		#tnfile = ofile + ".xml"
		#f = open(tid + ".xml", "w")
		f = open(tnfile, "w", encoding="utf-8")
		f.write(MainXML)
		f.close()
		
	
	return match_results

def MakeBatchFile(dcw_index):

	tbatch = ""
	for tid in dcw_index:
		tbatch += "REM \"Loading " + tid + ".xml" + "\"" + "\n"
		tbatch += "cernerload " + tid + ".xml" + "\n"
		
	f = open("load_orders.bat", "w")
	f.write(tbatch)
	f.close()

def ExtractOrderData(tfile, matchdata, match_results, ordermap, file_index):

	# Hierarchy: KNOWLEDGEPLAN => ORDERSETLIST => ORDERSET => CATEGORYLIST => 
	# COMPONENTLIST => COMPONENT => SENTENCELIST => SENTENCE
	knowledgeplan_list = defaultdict(str)
	phaselist = defaultdict(str)
	categorylist = defaultdict(lambda: defaultdict(str))
	component_type = defaultdict(lambda: defaultdict(lambda: defaultdict(str)))
	comp_precheck = defaultdict(lambda: defaultdict(lambda: defaultdict(str)))
	comp_required = defaultdict(lambda: defaultdict(lambda: defaultdict(str)))
	comp_index = defaultdict(lambda: defaultdict(str))
	componentlist = defaultdict(lambda: defaultdict(lambda: defaultdict(str)))
	sentencelist = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(str))))
	detaillist = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(str)))))
	categoryvocab = defaultdict(str)
	
	# 1- Plan Name
	# 2- Phase
	# 3- Clinical Category
	# 4- Sub Category
	# 5- Component Type
	# 6- Required (Orders) / Persistent (Notes)
	# 7- Prechecked - Can De-select
	# 8- Component
	# 9- IV Ingredient
	# 10- Order Sentence
	# 11- Order Comments
	# 12- First Sentence to Default
	# other: Offset	Offset Unit	Evidence Link	Dose	Dose Unit	Volume Dose	Volume Dose Unit
	# Rate	Rate Unit	Freetext Rate	Route	Frequency	PRN

	dcwcols = {"Plan Name": 1, "Phase": 2, "Clinical Category": 3, "Sub Category": 4, "Component Type": 5, "Required": 6, "Prechecked": 7, "Component": 8, "IV Ingredient": 9, "Order Sentence": 10, "Order Comments": 11}

	#tfile = "DCW Library - All PowerPlans in P0783 (DCW Format).xlsx"
	book = openpyxl.load_workbook(tfile)
	
	# get the right sheet - the only sheet, "MUM" sheet, or the active sheet 
	sheet = book.active
	numsheets = len(book.sheetnames)
	if numsheets == 1:
		sheet = book.active
	if numsheets > 1:
		tfound = False

		for tsheet in book.worksheets:
			ttitle = tsheet.title
			if ttitle.find('MUM') > -1:
				sheet = tsheet
				tfound = True

		#if tfound == False:
		#	for tsheet in book.worksheets:
		#		ttitle = tsheet.title
		#		if ttitle.find('Components') > -1:
		#			sheet = tsheet
		#			tfound = True
		
		if tfound == False:
			sheet = book.active
	
	row_count = sheet.max_row
	column_count = sheet.max_column
	#print("Current sheet: " + str(sheet.title))
	
	checkval = sheet.cell(row=3, column=1).value
	#print("Check value: " + checkval)
	
	if checkval == "Plan Name":

		# create json object
		orderdata = defaultdict(lambda: defaultdict(str))
		tplan_name = ""
		tphase = ""
		tcat = ""
		tsubcat = ""
		ttype = ""
		tcnt = 1
		orderdata = {}
		
		# get columns
		collist = []
		for x in range(1, column_count):
			tval = sheet.cell(row=3, column=x).value
			collist.append(tval)
			#print("Column: " + tval)

		# extract DCWs
		prevDCW = ""
		currDCW = ""
		currComponent = ""
		currCategory = ""
		dcwID = 0
		dcw_index = defaultdict(str)
		
		# get title
		ttitle = sheet.cell(row=4, column=1).value
		currDCW = ttitle
		dcw_index["0"] = currDCW
		prevDCW = currDCW
		
		for x in range(5, row_count):
		
			# get values
			catval = str(sheet.cell(row=x, column=3).value)
			tcat = catval
			
			# https://stackoverflow.com/questions/40464804/python-openpyxl-outputs-none-for-empty-cells
			if tcat == "None":
				tcat = ""
			
			if (tcat != "") and (tcat != "null"):
				categorylist[str(dcwID)][tcat] = '1'
				categoryvocab[tcat] = 1
				currCategory = tcat
			
			#print("Category: " + tcat)
			
			if (tcat == ""):
			
				tcompType = sheet.cell(row=x, column=5).value				
				if str(tcompType) != "":
					
					# remove unicode characters from certain fields
					#ttitle = sheet.cell(row=x, column=1).value
					tphase = str(sheet.cell(row=x, column=2).value).strip().encode("ascii", "ignore").decode()
					#tcat = sheet.cell(row=x, column=3).value
					tcompType = str(sheet.cell(row=x, column=5).value).strip().encode("ascii", "ignore").decode()
					#print("Component type: " + tcompType)
					treq = str(sheet.cell(row=x, column=6).value)
					tprecheck = str(sheet.cell(row=x, column=7).value)
					tcomponent = str(sheet.cell(row=x, column=8).value).strip().encode("ascii", "ignore").decode()
					tsentence = str(sheet.cell(row=x, column=10).value).strip().encode("ascii", "ignore").decode()
					
					# get details
					tevidence = str(sheet.cell(row=x, column=15).value)
					tdose = str(sheet.cell(row=x, column=16).value)
					tdoseunit = str(sheet.cell(row=x, column=17).value)
					tvolume = str(sheet.cell(row=x, column=18).value)
					tvolumeunit = str(sheet.cell(row=x, column=19).value)
					trate = str(sheet.cell(row=x, column=20).value)
					trateunit = str(sheet.cell(row=x, column=21).value)
					troute = str(sheet.cell(row=x, column=23).value)
					tfreq = str(sheet.cell(row=x, column=24).value)
					tprn = str(sheet.cell(row=x, column=25).value)
					
					#print("Component: " + tcomponent)
										
					if (tphase != "") and (tphase != "None"):
						phaselist[str(dcwID)] = tphase

					tcompFound = False
					if (tcomponent != "") and (tcomponent != "None"):
					
						if tcomponent in componentlist[str(dcwID)][currCategory]:
							tcompFound = True
					
						if tcompFound == False:
							componentlist[str(dcwID)][currCategory][tcomponent] = tcomponent
							currComponent = tcomponent
						
						if tcompFound == True:
							dupcomponent = tcomponent + " duplicate"
							dupindex = 2
							while dupcomponent in componentlist[str(dcwID)][currCategory]:
								dupcomponent = tcomponent + " duplicate " + str(dupindex)
								dupindex += 1
								
							componentlist[str(dcwID)][currCategory][dupcomponent] = dupcomponent
							currComponent = dupcomponent
						
					if (tcompType != "") and (tcompType != "None"):
						#component_type[str(dcwID)][currCategory][tcomponent] = tcompType
						component_type[str(dcwID)][currCategory][currComponent] = tcompType
						if tcompFound == True:
							component_type[str(dcwID)][currCategory][currComponent] = "Note"

					if (tprecheck != "") and (tprecheck != "None"):
						#comp_precheck[str(dcwID)][currCategory][tcomponent] = tprecheck
						comp_precheck[str(dcwID)][currCategory][currComponent] = tprecheck

					if (treq != "") and (treq != "None"):
						#comp_required[str(dcwID)][currCategory][tcomponent] = treq
						comp_precheck[str(dcwID)][currCategory][currComponent] = tprecheck

					if (tsentence != "") and (tsentence != "None"):
						sentencelist[str(dcwID)][currCategory][currComponent][tsentence] = tsentence

					# get details for each sentence
					if (tevidence != "") and (tevidence != "None"):
						detaillist[str(dcwID)][currCategory][currComponent][tsentence]['evidence'] = tevidence
					
					if (tdose != "") and (tdose != "None"):
						detaillist[str(dcwID)][currCategory][currComponent][tsentence]['dose'] = tdose
						
					if (tdoseunit != "") and (tdoseunit != "None"):
						detaillist[str(dcwID)][currCategory][currComponent][tsentence]['doseunit'] = tdoseunit

					if (tvolume != "") and (tvolume != "None"):
						detaillist[str(dcwID)][currCategory][currComponent][tsentence]['volume'] = tvolume

					if (tvolumeunit != "") and (tvolumeunit != "None"):
						detaillist[str(dcwID)][currCategory][currComponent][tsentence]['volumeunit'] = tvolumeunit

					if (trate != "") and (trate != "None"):
						detaillist[str(dcwID)][currCategory][currComponent][tsentence]['rate'] = trate

					if (trateunit != "") and (trateunit != "None"):
						detaillist[str(dcwID)][currCategory][currComponent][tsentence]['rateunit'] = trateunit

					if (troute != "") and (troute != "None"):
						detaillist[str(dcwID)][currCategory][currComponent][tsentence]['route'] = troute

					if (tfreq != "") and (tfreq != "None"):
						detaillist[str(dcwID)][currCategory][currComponent][tsentence]['frequency'] = tfreq

					if (tprn != "") and (tprn != "None"):
						detaillist[str(dcwID)][currCategory][currComponent][tsentence]['prn'] = tprn

		# Create XML Files from parsed data
		#if tfile.find("PED ED Acetaminophen (Tylenol) Overdose NAC.xlsx") > -1:
		#	print(str(componentlist))
		match_results = XMLMakeFiles(tfile, dcw_index, phaselist, categorylist, component_type, comp_precheck, comp_required, componentlist, sentencelist, detaillist, matchdata, match_results, ordermap, file_index)
		
		# make batch loader
		#MakeBatchFile(dcw_index)
		
		# write out current categories
		#for tcat in categoryvocab:
		#	print("Category: " + tcat)

		'''
		# export html
		dcwID = 1
		currDCW = ""
		prevDCW = ""
		ttable = []
		dcw_index = defaultdict(str)
		for x in range(4, row_count):
			tval = sheet.cell(row=x, column=1).value
			if tval != "":
				if prevDCW != "":
					MakeHTMLFile(dcwID, currDCW, collist, ttable)				
					ttable = []
				prevDCW = currDCW
				currDCW = tval
				dcwID += 1
				print(currDCW)
				dcw_index[str(dcwID)] = currDCW
				
			currRow = []		
			for y in range(1, column_count):
				tval = sheet.cell(row=x, column=y).value
				currRow.append(str(tval))

			ttable.append(currRow)
		
		# create JSON index
		#MakeJSONIndex(dcw_index)
		'''
		
		'''
		for x in range(4, row_count):
			tval = sheet.cell(row=x, column=1).value
			if tval != "":
				tplan_name = str(tval)
				#orderdata[tplan_name]["name"] = tval
			tval = sheet.cell(row=x, column=2).value
			if tval != "":
				tphase = tval
				#orderdata[tplan_name]["phase"] = tval
			tval = sheet.cell(row=x, column=3).value
			if tval != "":
				tcategory = tval
				#orderdata[tplan_name]["category"] = tval
			tval = sheet.cell(row=x, column=5).value
			if tval != "":
				ttype = tval
				#orderdata[tplan_name]["component_type"] += tval + "|"
			tval = sheet.cell(row=x, column=6).value
						
			#print(tplan_name + "\t" + str(tphase) + "\t" + str(tcat) + "\t" + str(tsubcat) + "\t" + str(ttype) + "\n")
		'''	

	return match_results

# get current global version
mainversion = GetCurrentVersion()

# Get a list of DCWs
#tbatch = "Batch1"
#tbatch = "Batch2"
#tbatch = "March3"
#tbatch = "March6"
#tbatch = "First100"
#tbatch = "Last30"
tbatch = "FixBatch1"
cst_path = "C:\\Apache24\\htdocs\\orders\\data\\high_priority\\" + tbatch
filelist = GetCSTList(cst_path)
#filelist = LoadPriorityCSTList("high_priority.tsv")
#filelist = LoadPriorityCSTList("neph_orders.tsv")

#filelist = ["c:\\Apache24\\htdocs\\orders\\data\\high_priority\\DCWs\\CARD Heart Failure.xlsx"]

matchdata = LoadNHOrderData()
match_results = defaultdict(str)

#ordermap = LoadNHToCSTMap()
ordermap = LoadCurrentMapping()

# parse excel sheet into JSON
tcnt = 0
for x in range(0, len(filelist)):
	#if tcnt < 3:
	nfile = filelist[x]
	ofiledata = nfile.split('\\')
	ofile = ofiledata[len(ofiledata) - 1]
	print("Processing: " + ofile)
	#if tcnt < 10:
	match_results = ExtractOrderData(nfile, matchdata, match_results, ordermap, x)
	tcnt += 1

print("matching:")
print(str(match_results))

# write out mapping between NH order library and CST
f = open("nh_to_cst_map_" + tbatch + ".tsv", "w")
for tid in match_results:
	tresult = match_results[tid].strip().lower()
	tmatch = ordermap[tid].strip()
	if (tresult == "no nh match"):
		f.write(tid + "\t" + tresult + "\n")
f.close()

# write XML
#data = defaultdict(str)
#xmlorderprop = GetOrderSetProp()
#CreateXMLDocument(xmlorderprop, data, "test.xml")
