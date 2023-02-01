import openpyxl
from xml.dom import minidom
from collections import defaultdict

def PrintLine(trow):

	tline = ""
	for x in range(0, len(trow)):
		tline += trow[x] + "\t"
	tline = tline.strip() + "\n"
	
	return tline
	
def GetIndent(tlevel):

	tline = ""
	for x in range(0, tlevel):
		tline += "\t"
	
	return tline

def PrintHTMLLine(trow):

	tline = "<tr>"
	for x in range(0, len(trow)):
		tline += "<td>" + trow[x] + "</td>"
	tline = tline.strip() + "</tr>\n"
	
	return tline


def MakeHTMLFile(tid, ttitle, collist, ttable):

	tfile = str(tid) + ".html"
	f = open(tfile, "w")
	f.write("<html><head><title>" + ttitle + "</title>")
	f.write('<link rel="stylesheet" href="cst_table.css">' + "\n")
	f.write("</head><body style=\"font-family: arial;\">")
	f.write("<h2>" + ttitle + "</h2>\n")
	f.write("<table id=\"customers\">\n")
	theader = PrintHTMLLine(collist)
	f.write(theader)
	for trow in ttable:
		tline = PrintHTMLLine(trow)
		f.write(tline)
	f.write("</table></body></html>\n")
	f.close()

def MakeJSONIndex(datatable):

	tjson = '[\n'
	for tid in datatable:
		tjson += "\t" + '{"dcw_id": "' + tid + '", "title": "' + datatable[tid] + '"},' + "\n"
	tjson = tjson[:-2] + "\n"
	tjson += ']\n'

	f = open("cst_dcws.json", "w")
	f.write(tjson)
	f.close()

def GetOrderSetProp():

	txml = ""
	with open("head_xml_data.dat", "r") as infile:
		for line in infile:
			line = line.strip()
			txml = txml + "\t\t" + line + "\n"
	infile.close()
	
	return txml
	
def CreateHTMLPages():

	tfile = "DCW Library - All PowerPlans in P0783 (DCW Format).xlsx"
	book = openpyxl.load_workbook(tfile)
	sheet = book.active 
	row_count = sheet.max_row
	column_count = sheet.max_column

	orderdata = defaultdict(lambda: defaultdict(str))
	tplan_name = ""
	tphase = ""
	tcat = ""
	tsubcat = ""
	ttype = ""
	tcnt = 1
	orderdata = {}	
	for x in range(1, row_count):
		tval = sheet.cell(row=x, column=1).value
		if tval != "":
			orderdata[tplan_name]["name"] = tval	

def GetNHCategory(tval):

	# NH Categories on B1978 from Tyler Scott (Dec 20, 2022)
	# NH Category                   CST Category
	# -----------                   ------------
	# Admit/Transfer/Discharge      Admit/Transfer/Discharge
	# Patient Care                  Patient Care
	# Activity                      Activity
	# Diet/Nutrition                Diet/Nutrition
	# Continuous Infusions          Continuous Infusions
	# Medications                   Medications
	# Laboratory                    Laboratory
	# Diagnostic Tests              Diagnostic Tests
	# Respiratory                   Respiratory
	# Allied Health                 Allied Health
	# Consults/Referrals            Consults/Referrals
	# Communication Orders          Communication Orders
	# Procedures                    Procedures
	# Non Categorized               -----
	# Allergies                     -----
	# Diagnoses                     -----
	# Medical Supplies              Supplies
	# Special Procedures            -----
	# Other Test                    -----
	# Supplies                      Supplies
	
	retval = tval + "|Clinical Categories"

	return retval

def XMLBuildDetailList(dcwid, catid, compid, sentid, detail_list):

	# count up details
	tXML = ""
	tcnt = 0
	for tid in detail_list[dcwid][catid][compid][sentid]:
		tcnt += 1
	
	if tcnt > 0:
		tXML = "\n" + GetIndent(9) + "<DETAILLIST>\n"
		for tid in detail_list[dcwid][catid][compid][sentid]:
			tXML += GetIndent(10) + "<DETAILS>\n"
			tXML += GetIndent(11) + "<FIELDMEAN>" + tid + "</FIELDMEAN>\n"
			tXML += GetIndent(11) + "<FIELDDESC>" + tid + "</FIELDDESC>\n"
			tXML += GetIndent(11) + "<FIELDDISPVALUE>" + str(detail_list[dcwid][catid][compid][sentid][tid]) + "</FIELDDISPVALUE>\n"
			tXML += GetIndent(10) + "</DETAILS>\n"
		tXML += GetIndent(9) + "</DETAILLIST>\n"

	return tXML

def XMLBuildSentenceList(dcwid, catid, compid, sentence_list, detail_list):

	# count up sentences
	tXML = ""
	tcnt = 0
	for tid in sentence_list[dcwid][catid][compid]:
		tcnt += 1
	
	if tcnt > 0:
		tXML = "\n" + GetIndent(7) + "<SENTENCELIST>\n"
		for tid in sentence_list[dcwid][catid][compid]:
			tXML += GetIndent(8) + "<SENTENCE>\n"
			tXML += GetIndent(9) + "<DISPLAYLINE>" + sentence_list[dcwid][catid][compid][tid] + "</DISPLAYLINE>\n"
			tXML += GetIndent(9) + XMLBuildDetailList(dcwid, catid, compid, tid, detail_list)
			tXML += "\n" + GetIndent(8) + "</SENTENCE>\n"
		tXML += GetIndent(7) + "</SENTENCELIST>\n"

	return tXML
	
def XMLBuildComponentList(dcwid, catid, component_list, sentence_list, detail_list):

	tXML = "\n" + GetIndent(5) + "<COMPONENTLIST>\n"
	for tid in component_list[dcwid][catid]:
		tXML += GetIndent(6) + "<COMPONENT>\n"
		tXML += GetIndent(7) + "<CAPTION>" + component_list[dcwid][catid][tid] + "</CAPTION>\n"
		tXML += GetIndent(7) + "<CKI/>\n"
		tXML += GetIndent(7) + "<DEFAULTOSIND>1</DEFAULTOSIND>\n"
		tXML += GetIndent(7) + "<DEFAULTSELECTED>1</DEFAULTSELECTED>\n"
		tXML += GetIndent(7) + "<ORDERABLETYPE>O</ORDERABLETYPE>\n"
		#print(dcwid + " - Sentences: " + str(sentence_list[dcwid][catid][tid]))
		tXML += GetIndent(7) + XMLBuildSentenceList(dcwid, catid, tid, sentence_list, detail_list)
		tXML += GetIndent(6) + "</COMPONENT>\n"
	tXML += GetIndent(5) + "</COMPONENTLIST>\n"

	return tXML

def XMLBuildCategoryList(dcwid, category_list, component_list, sentence_list, detail_list):

	tXML = GetIndent(3) + "<CATEGORYLIST>\n"
	for tid in category_list[dcwid]:
	
		# get NH category
		ncat = GetNHCategory(tid)
	
		tXML += GetIndent(4) + "<CATEGORY>\n"
		tXML += GetIndent(5) + "<CAPTION>" + tid + "</CAPTION>\n"
		tXML += GetIndent(5) + "<CLINICALCATEGORYMEAN>" + ncat + "</CLINICALCATEGORYMEAN>\n"
		#tXML += GetIndent(3) + "<meaning>" + category_list[tid] + "</meaning>\n"
		tXML += GetIndent(3) + XMLBuildComponentList(dcwid, tid, component_list, sentence_list, detail_list)
		tXML += GetIndent(4) + "</CATEGORY>\n"
	tXML += GetIndent(3) + "</CATEGORYLIST>\n"

	return tXML

def XMLMakeRoot(ttitle, tphase):

	tXML = '<?xml version="1.0" encoding="UTF-8"?>' + "\n"
	tXML += "<KNOWLEDGEPLAN>\n"
	tXML += "\t" + "<SOURCEORDERSETTYPE>CAREPLAN</SOURCEORDERSETTYPE>\n"
	tXML += "\t" + "<CAPTION>" + ttitle + "</CAPTION>\n"
	tXML += "\t" + "<ORDERSETLIST>\n"
	tXML += "\t\t" + "<ORDERSET>\n"
	tXML += "\t\t\t" + "<CAPTION>" + ttitle + "</CAPTION>\n"
	tXML += "\t\t\t" + "<DISPLAY>" + ttitle + "</DISPLAY>\n"
	tXML += "\t\t\t" + "<EVIDENCEURL/>" + "\n"
	tXML += "\t\t\t" + "<EVIDENCEURLTYPEMEAN/>" + "\n"
	tXML += "\t\t\t" + "<CKI/>" + "\n"
	tXML += "\t\t\t" + "<VERSION>1</VERSION>" + "\n"
	tXML += "\t\t\t" + "<DURATION>0</DURATION>" + "\n"
	tXML += "\t\t\t" + "<DURATIONUNITMEAN/>" + "\n"
	tXML += "\t\t\t" + "<SUBPHASEIND>0</SUBPHASEIND>" + "\n"
	tXML += "\t\t\t" + "<STANDARDCYCLENBR>0</STANDARDCYCLENBR>" + "\n"
	tXML += "\t\t\t" + "<CYCLEIND>0</CYCLEIND>" + "\n"
	tXML += "\t\t\t" + "<DIAGNOSISCAPTUREIND>0</DIAGNOSISCAPTUREIND>" + "\n"
	tXML += "\t\t\t" + "<PROVIDERPROMPTIND>0</PROVIDERPROMPTIND>" + "\n"
	tXML += "\t\t\t" + "<HIDEFLEXEDCOMPIND>0</HIDEFLEXEDCOMPIND>" + "\n"
	tXML += "\t\t\t" + "<DEFAULTVIEWMEAN/>" + "\n"
	tXML += "\t\t\t" + "<ALLOWCOPYFORWARDIND>0</ALLOWCOPYFORWARDIND>" + "\n"
	tXML += "\t\t\t" + "<AUTOINITIATEIND>0</AUTOINITIATEIND>" + "\n"
	tXML += "\t\t\t" + "<ALERTSONPLANIND>0</ALERTSONPLANIND>" + "\n"
	tXML += "\t\t\t" + "<ALERTSONPLANUPDIND>0</ALERTSONPLANUPDIND>" + "\n"
	tXML += "\t\t\t" + "<PROMPTONSELECTIONIND>0</PROMPTONSELECTIONIND>" + "\n"
	tXML += "\t\t\t" + "<DEFAULTVISITTYPEFLAG>0</DEFAULTVISITTYPEFLAG>" + "\n"
	tXML += "\t\t\t" + "<CYCLEBEGINNBR>0</CYCLEBEGINNBR>" + "\n"
	tXML += "\t\t\t" + "<CYCLEENDNBR>0</CYCLEENDNBR>" + "\n"
	tXML += "\t\t\t" + "<CYCLELABELDISP/>" + "\n"
	tXML += "\t\t\t" + "<CYCLELOCKENDIND>0</CYCLELOCKENDIND>" + "\n"
	tXML += "\t\t\t" + "<CYCLEDISPLAYENDIND>0</CYCLEDISPLAYENDIND>" + "\n"
	tXML += "\t\t\t" + "<CYCLEINCREMENTNBR>0</CYCLEINCREMENTNBR>" + "\n"
	tXML += "\t\t\t" + "<DEFAULTACTIONINFUTUREMEAN/>" + "\n"
	tXML += "\t\t\t" + "<DEFAULTACTIONINNOWMEAN/>" + "\n"
	tXML += "\t\t\t" + "<DEFAULTACTIONOUTFUTUREMEAN/>" + "\n"
	tXML += "\t\t\t" + "<DEFAULTACTIONOUTNOWMEAN/>" + "\n"
	tXML += "\t\t\t" + "<FUTUREIND>0</FUTUREIND>" + "\n"
	tXML += "\t\t\t" + "<OPTIONALIND>0</OPTIONALIND>" + "\n"
	tXML += "\t\t\t" + "<PATHWAYCLASSMEAN/>" + "\n"
	tXML += "\t\t\t" + "<ROUTEFORREVIEWIND>0</ROUTEFORREVIEWIND>" + "\n"
	tXML += "\t\t\t" + "<DEFAULTSTARTTIMETXT/>" + "\n"
	tXML += "\t\t\t" + "<PRIMARYIND>0</PRIMARYIND>" + "\n"
	tXML += "\t\t\t" + "<PERIODNBR>0</PERIODNBR>" + "\n"
	tXML += "\t\t\t" + "<PERIODCUSTOMLABEL/>" + "\n"
	tXML += "\t\t\t" + "<RESCHEDULEREASONACCEPTFLAG>0</RESCHEDULEREASONACCEPTFLAG>" + "\n"
	tXML += "\t\t\t" + "<OPENBYDEFAULTIND>0</OPENBYDEFAULTIND>" + "\n"
	tXML += "\t\t\t" + "<ALLOWACTIVATEALLIND>1</ALLOWACTIVATEALLIND>" + "\n"
	tXML += "\t\t\t" + "<REVIEWREQUIREDSIGCOUNT>0</REVIEWREQUIREDSIGCOUNT>" + "\n"
	tXML += "\t\t\t" + "<RESTRICTEDACTIONSBITMASK>0</RESTRICTEDACTIONSBITMASK>" + "\n"
	tXML += "\t\t\t" + "<OVERRIDEMRDONPLANIND>0</OVERRIDEMRDONPLANIND>" + "\n"
	tXML += "\t\t\t" + "<LINKEDPHASEIND>0</LINKEDPHASEIND>" + "\n"

	tXML += '{%CategoryList%}'

	tXML += "\n\t\t" + "</ORDERSET>\n"	
	tXML += "\t" + "</ORDERSETLIST>\n"
	tXML += "</KNOWLEDGEPLAN>\n"

	return tXML
	
def XMLMakeFiles(dcw_index, phaselist, categorylist, component_type, comp_precheck, comp_required, componentlist, sentencelist, detaillist):

	# Create the XML Document Stub and Components
	# Using C:\Apache24\htdocs\orders\data\island_xml\Subcutaneous Insulin Cerner PowePlan V3.xml
	for tid in dcw_index:
		ttitle = dcw_index[tid]
		tphase = phaselist[tid]
		
		# Create XML Root Code
		MainXML = XMLMakeRoot(ttitle, tphase)
		
		# Create Category Data
		CategoryXML = XMLBuildCategoryList(tid, categorylist, componentlist, sentencelist, detaillist)
		MainXML = MainXML.replace('{%CategoryList%}', CategoryXML)
	
		# Create XML Components and Sentences
		f = open(tid + ".xml", "w")
		f.write(MainXML)
		f.close()

def MakeBatchFile(dcw_index):

	tbatch = ""
	for tid in dcw_index:
		tbatch += "REM \"Loading " + tid + ".xml" + "\"" + "\n"
		tbatch += "cernerload " + tid + ".xml" + "\n"
		
	f = open("load_orders.bat", "w")
	f.write(tbatch)
	f.close()

def ExtractOrderData():

	# Hierarchy: KNOWLEDGEPLAN => ORDERSETLIST => ORDERSET => CATEGORYLIST => 
	# COMPONENTLIST => COMPONENT => SENTENCELIST => SENTENCE
	knowledgeplan_list = defaultdict(str)
	phaselist = defaultdict(str)
	categorylist = defaultdict(lambda: defaultdict(str))
	component_type = defaultdict(lambda: defaultdict(lambda: defaultdict(str)))
	comp_precheck = defaultdict(lambda: defaultdict(lambda: defaultdict(str)))
	comp_required = defaultdict(lambda: defaultdict(lambda: defaultdict(str)))
	comp_index = defaultdict(lambda: defaultdict(str))
	componentlist = defaultdict(lambda: defaultdict(lambda: defaultdict(str)))
	sentencelist = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(str))))
	detaillist = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(str)))))
	categoryvocab = defaultdict(str)
	
	# 1- Plan Name
	# 2- Phase
	# 3- Clinical Category
	# 4- Sub Category
	# 5- Component Type
	# 6- Required (Orders) / Persistent (Notes)
	# 7- Prechecked - Can De-select
	# 8- Component
	# 9- IV Ingredient
	# 10- Order Sentence
	# 11- Order Comments
	# 12- First Sentence to Default
	# other: Offset	Offset Unit	Evidence Link	Dose	Dose Unit	Volume Dose	Volume Dose Unit
	# Rate	Rate Unit	Freetext Rate	Route	Frequency	PRN

	dcwcols = {"Plan Name": 1, "Phase": 2, "Clinical Category": 3, "Sub Category": 4, "Component Type": 5, "Required": 6, "Prechecked": 7, "Component": 8, "IV Ingredient": 9, "Order Sentence": 10, "Order Comments": 11}

	tfile = "DCW Library - All PowerPlans in P0783 (DCW Format).xlsx"
	book = openpyxl.load_workbook(tfile)
	sheet = book.active 
	row_count = sheet.max_row
	column_count = sheet.max_column

	# create json object
	orderdata = defaultdict(lambda: defaultdict(str))
	tplan_name = ""
	tphase = ""
	tcat = ""
	tsubcat = ""
	ttype = ""
	tcnt = 1
	orderdata = {}
	
	# get columns
	collist = []
	for x in range(1, column_count):
		tval = sheet.cell(row=3, column=x).value
		collist.append(tval)
		#print("Column: " + tval)

	# extract DCWs
	prevDCW = ""
	currDCW = ""
	currComponent = ""
	currCategory = ""
	dcwID = 0
	dcw_index = defaultdict(str)
	for x in range(4, row_count):
	
		# get values
		ttitle = sheet.cell(row=x, column=1).value
		tphase = sheet.cell(row=x, column=2).value
		tcat = sheet.cell(row=x, column=3).value
		tcompType = sheet.cell(row=x, column=5).value
		treq = sheet.cell(row=x, column=6).value
		tprecheck = sheet.cell(row=x, column=7).value
		tcomponent = sheet.cell(row=x, column=8).value
		tsentence = sheet.cell(row=x, column=10).value
		
		# get details
		tevidence = sheet.cell(row=x, column=15).value
		tdose = sheet.cell(row=x, column=16).value
		tdoseunit = sheet.cell(row=x, column=17).value
		tvolume = sheet.cell(row=x, column=18).value
		tvolumeunit = sheet.cell(row=x, column=19).value
		trate = sheet.cell(row=x, column=20).value
		trateunit = sheet.cell(row=x, column=21).value
		troute = sheet.cell(row=x, column=23).value
		tfreq = sheet.cell(row=x, column=24).value
		tprn = sheet.cell(row=x, column=25).value
		
		if ttitle != "":
			dcwID += 1
			currDCW = ttitle
			#print(currDCW)
			dcw_index[str(dcwID)] = currDCW
			prevDCW = currDCW

		if tphase != "":
			phaselist[str(dcwID)] = tphase

		if tcat != "":
			categorylist[str(dcwID)][tcat] = '1'
			categoryvocab[tcat] = 1
			currCategory = tcat

		if tcomponent != "":
			#print("Found component: " + tcomponent)
			componentlist[str(dcwID)][currCategory][tcomponent] = tcomponent
			currComponent = tcomponent

		if tcompType != "":
			component_type[str(dcwID)][currCategory][tcomponent] = tcompType

		if tprecheck != "":
			comp_precheck[str(dcwID)][currCategory][tcomponent] = tprecheck

		if treq != "":
			comp_required[str(dcwID)][currCategory][tcomponent] = treq

		if tsentence != "":
			sentencelist[str(dcwID)][currCategory][currComponent][tsentence] = tsentence

		# get details for each sentence
		if tevidence != "":
			detaillist[str(dcwID)][currCategory][currComponent][tsentence]['evidence'] = tevidence
		
		if tdose != "":
			detaillist[str(dcwID)][currCategory][currComponent][tsentence]['dose'] = tdose
			
		if tdoseunit != "":
			detaillist[str(dcwID)][currCategory][currComponent][tsentence]['doseunit'] = tdoseunit

		if tvolume != "":
			detaillist[str(dcwID)][currCategory][currComponent][tsentence]['volume'] = tvolume

		if tvolumeunit != "":
			detaillist[str(dcwID)][currCategory][currComponent][tsentence]['volumeunit'] = tvolumeunit

		if trate != "":
			detaillist[str(dcwID)][currCategory][currComponent][tsentence]['rate'] = trate

		if trateunit != "":
			detaillist[str(dcwID)][currCategory][currComponent][tsentence]['rateunit'] = trateunit

		if troute != "":
			detaillist[str(dcwID)][currCategory][currComponent][tsentence]['route'] = troute

		if tfreq != "":
			detaillist[str(dcwID)][currCategory][currComponent][tsentence]['frequency'] = tfreq

		if tprn != "":
			detaillist[str(dcwID)][currCategory][currComponent][tsentence]['prn'] = tprn

	# Create XML Files from parsed data
	XMLMakeFiles(dcw_index, phaselist, categorylist, component_type, comp_precheck, comp_required, componentlist, sentencelist, detaillist)
	
	# make batch loader
	MakeBatchFile(dcw_index)
	
	# write out current categories
	for tcat in categoryvocab:
		print("Category: " + tcat)

	'''
	# export html
	dcwID = 1
	currDCW = ""
	prevDCW = ""
	ttable = []
	dcw_index = defaultdict(str)
	for x in range(4, row_count):
		tval = sheet.cell(row=x, column=1).value
		if tval != "":
			if prevDCW != "":
				MakeHTMLFile(dcwID, currDCW, collist, ttable)				
				ttable = []
			prevDCW = currDCW
			currDCW = tval
			dcwID += 1
			print(currDCW)
			dcw_index[str(dcwID)] = currDCW
			
		currRow = []		
		for y in range(1, column_count):
			tval = sheet.cell(row=x, column=y).value
			currRow.append(str(tval))

		ttable.append(currRow)
	
	# create JSON index
	#MakeJSONIndex(dcw_index)
	'''
	
	'''
	for x in range(4, row_count):
		tval = sheet.cell(row=x, column=1).value
		if tval != "":
			tplan_name = str(tval)
			#orderdata[tplan_name]["name"] = tval
		tval = sheet.cell(row=x, column=2).value
		if tval != "":
			tphase = tval
			#orderdata[tplan_name]["phase"] = tval
		tval = sheet.cell(row=x, column=3).value
		if tval != "":
			tcategory = tval
			#orderdata[tplan_name]["category"] = tval
		tval = sheet.cell(row=x, column=5).value
		if tval != "":
			ttype = tval
			#orderdata[tplan_name]["component_type"] += tval + "|"
		tval = sheet.cell(row=x, column=6).value
					
		#print(tplan_name + "\t" + str(tphase) + "\t" + str(tcat) + "\t" + str(tsubcat) + "\t" + str(ttype) + "\n")
	'''	

# parse excel sheet into JSON
ExtractOrderData()

# write XML
#data = defaultdict(str)
#xmlorderprop = GetOrderSetProp()
#CreateXMLDocument(xmlorderprop, data, "test.xml")

# Zynx Import Log - Jan 3, 2023
# Order sets with errors:

#7
#10
#11
#12
#15
#18
#29
#35
#39
#43
